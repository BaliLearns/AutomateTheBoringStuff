#!python3
#multiplicationTable.py -

'''

Multiplication Table Maker
Create a program multiplicationTable.py that takes a number N from the command line and
creates an N×N multiplication table in an Excel spreadsheet. For example, when the program
is run like this:

Row 1 and column A should be used for labels and should be in bold.

'''

import openpyxl
from openpyxl.styles import Font
import os, sys
from openpyxl.utils import get_column_letter

#Simple checks for arguments passed
if len(sys.argv) > 2:
    raise Exception('More than 1 arguments passed')
try:
    number = int(sys.argv[1])
except ValueError:
    print('Input in not in integer format')
    sys.exit()

#Workbook creation
wb = openpyxl.Workbook()
sheet = wb.active
fontstyle = Font(bold = True)

#Formatting of 1st row and column
for num in range(1, number+1):
    sheet.cell(row = 1, column = num+1).value = num
    sheet.cell(row = 1, column = num+1).font = fontstyle
    sheet.cell(row = num+1, column = 1).value = num
    sheet.cell(row = num+1, column = 1).font = fontstyle

#changing the values of cells in between to formulas
for row in range(2, number+2):
    for col in range(2, number+2):
        sheet.cell(row = row, column = col).value = '=A'+ str(row) + '*'+ get_column_letter(col) + str(1)

#saving the file
wb.save('multiplicationTable' + str(number) + '.xlsx')
